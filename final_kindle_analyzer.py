
"""
Final Kindle Log Analyzer with PDF Export, Enhanced TXT, and Waveform Boxes
"""
import sys
import os
import re
from datetime import datetime
from pathlib import Path
# Import the new export and visualization modules
try:
    from pdf_export import PdfExporter
    from txt_export import TxtExporter
    from waveform_plot import WaveformVisualizer
    ENHANCED_EXPORTS_AVAILABLE = True
except ImportError as e:
    print(f"Warning: Enhanced export modules not available: {e}")
    ENHANCED_EXPORTS_AVAILABLE = False

from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
    QHBoxLayout, QTextEdit, QPushButton, QLabel, 
    QTableWidget, QTableWidgetItem, QTabWidget,
    QSplitter, QGroupBox, QFileDialog, QProgressBar, 
    QLineEdit, QComboBox, QListWidget, QMessageBox,
    QHeaderView, QAbstractItemView, QCheckBox, QGridLayout,
    QFrame, QScrollArea)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer
from PyQt5.QtGui import QFont, QColor, QBrush, QPalette, QPixmap
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

class FixedSuspendEventParser:
    """Fixed parser for suspend mode (Power Button) based on user's sample"""
    
    def extract_marker(self, line):
        """Extract marker number from log line"""
        match1 = re.search(r'EPDC\]\[(\d+)\]', line)
        if match1:
            return match1.group(1)
        
        match2 = re.search(r'mxc_epdc_fb: \[(\d+)\]', line)
        if match2:
            return match2.group(1)
        
        return None

    def extract_height_and_waveform(self, line):
        """Extract height and waveform information from log line"""
        height_match = re.search(r'height=(\d+)', line)
        if not height_match:
            height_match = re.search(r'width=\d+, height=(\d+)', line)
        
        waveform_patterns = [
            r'new waveform = (?:0x)?[\da-f]+ \(([\w_() ]+)\)', 
            r'waveform:(?:0x)?[\da-f]+ \(([\w_() ]+)\)',
            r'waveform=(?:0x)?[\da-f]+ \(([\w_() ]+)\)',
            r'Sending update\. waveform:(?:0x)?[\da-f]+ \(([\w_() ]+)\)'
        ]
        
        waveform_name = None
        for pattern in waveform_patterns:
            match = re.search(pattern, line)
            if match:
                waveform_name = match.group(1).strip()
                break
        
        if height_match:
            height = int(height_match.group(1))
            return {
                'height': height,
                'waveform': waveform_name if waveform_name else "unknown"
            }
        return None

    def extract_end_timestamp(self, line):
        """Extract end timestamp from log line"""
        match = re.search(r'end time=(\d+)', line)
        if match:
            timestamp_str = match.group(1)
            last_6 = timestamp_str[-6:]
            return int(last_6)
        return None
    
    def extract_start_timestamp(self, line):
        """Extract start timestamp from power button press - FIXED VERSION"""
        # Look for the specific pattern: def:pbpress:time=XXXX.XXX:Power button pressed
        match = re.search(r'def:pbpress:time=(\d+\.\d+):Power button pressed', line)
        if match:
            timestamp_str = match.group(1)
            parts = timestamp_str.split('.')
            if len(parts) == 2:
                # Take last 3-6 digits before dot, then first 3 digits after dot
                before_dot = parts[0]
                after_dot = parts[1][:3]  # First 3 digits after dot
                
                # Take last 3 digits before dot to match the format
                last_digits_before = before_dot[-3:]
                result = int(last_digits_before + after_dot)
                return result
        
        return None

class LogProcessor(QThread):
    """Enhanced log processor with original log storage"""
    progress_updated = pyqtSignal(int)
    result_ready = pyqtSignal(dict)
    error_occurred = pyqtSignal(str)
    
    def __init__(self, log_content, mode="default"):
        super().__init__()
        self.log_content = log_content
        self.mode = mode
    
    def run(self):
        try:
            self.progress_updated.emit(10)
            
            # Split content into iterations while preserving original content
            iterations = re.split(r'ITERATION_(\d+)', self.log_content)[1:]
            
            if not iterations:
                iterations = ["01", self.log_content]
            
            self.progress_updated.emit(30)
            
            # Pair iteration numbers with content
            iteration_pairs = []
            for i in range(0, len(iterations), 2):
                if i+1 < len(iterations):
                    iteration_num = iterations[i]
                    iteration_content = iterations[i+1]
                    iteration_pairs.append((iteration_num, iteration_content))
            
            self.progress_updated.emit(50)
            
            results = []
            total_iterations = len(iteration_pairs)
            
            for idx, (iteration_num, iteration_content) in enumerate(iteration_pairs):
                lines = iteration_content.split('\n')
                result = self.process_iteration(lines, iteration_num, self.mode)
                
                if result:
                    # Store original log content with the result
                    result['original_log'] = iteration_content.strip()
                    results.append(result)
                
                progress = 50 + (idx + 1) * 40 // total_iterations
                self.progress_updated.emit(progress)
            
            self.progress_updated.emit(100)
            self.result_ready.emit({'results': results, 'total_iterations': len(iteration_pairs)})
            
        except Exception as e:
            self.error_occurred.emit(str(e))
    
    def process_iteration(self, lines, iteration_num, mode="default"):
        """Process a single iteration with fixed suspend parsing"""
        
        # Get appropriate parser
        if mode == "suspend":
            parser = FixedSuspendEventParser()
        else:
            # Use original parsers for default and swipe
            if mode == "default":
                parser = DefaultEventParser()
            else:  # swipe
                parser = SwipeEventParser()
        
        start_time = None
        start_line = None  # Store the line that contains start time for highlighting
        end_times_by_marker = {}
        heights_by_marker = {}
        current_marker = None
        
        for line in lines:
            if not line.strip():
                continue
            
            # Check for start time based on the parser's implementation
            if not start_time:
                possible_start = parser.extract_start_timestamp(line)
                if possible_start:
                    start_time = possible_start
                    start_line = line.strip()  # Store the start line for highlighting
            
            marker = parser.extract_marker(line)
            if marker:
                current_marker = marker
            
            if "Sending update" in line and current_marker:
                height_waveform = parser.extract_height_and_waveform(line)
                if height_waveform:
                    height = height_waveform['height']
                    waveform = height_waveform['waveform']
                    
                    heights_by_marker[current_marker] = {
                        'height': height,
                        'waveform': waveform if waveform and waveform != "auto" else "unknown",
                        'line': line.strip()
                    }
            
            if "update end marker=" in line and "end time=" in line:
                end_marker_match = re.search(r'update end marker=(\d+)', line)
                if end_marker_match:
                    end_marker = end_marker_match.group(1)
                    end_time = parser.extract_end_timestamp(line)
                    if end_time:
                        end_times_by_marker[end_marker] = {
                            'time': end_time,
                            'line': line.strip()
                        }
        
        if not start_time or not heights_by_marker or not end_times_by_marker:
            return None
        
        # Filter out markers with "unknown" waveforms
        valid_heights = {
            marker: info for marker, info in heights_by_marker.items() 
            if info['waveform'].lower() != "unknown"
        }
        
        # If no valid waveforms found, use all heights as a fallback
        if not valid_heights:
            valid_heights = heights_by_marker
        
        # Find the marker with the maximum height among valid waveforms
        max_height = max(info['height'] for info in valid_heights.values())
        max_height_markers = [marker for marker, info in valid_heights.items() 
                             if info['height'] == max_height]
        
        max_height_markers.sort(key=lambda m: int(m) if m.isdigit() else 0)
        chosen_marker = max_height_markers[-1] if max_height_markers else list(valid_heights.keys())[0]
        
        max_height_info = valid_heights[chosen_marker]
        
        # Get the end time for the chosen marker
        if chosen_marker in end_times_by_marker:
            max_height_end_time = end_times_by_marker[chosen_marker]['time']
        else:
            # If no end time for the chosen marker, use the maximum end time
            if end_times_by_marker:
                max_height_end_time = max(end_times_by_marker.values(), key=lambda x: x['time'])['time']
            else:
                return None
        
        # Calculate duration
        duration = max_height_end_time - start_time
        if duration < 0:
            duration = abs(duration)

        # Convert duration from milliseconds to seconds
        duration = duration / 1000.0

        return {
            'iteration': int(iteration_num),  # Convert to integer here
            'start': start_time,
            'stop': max_height_end_time,
            'marker': chosen_marker,
            'duration': duration,
            'max_height': max_height_info['height'],
            'max_height_waveform': max_height_info['waveform'],
            'start_line': start_line,  # For PDF highlighting
            'all_heights': [{'marker': m, 'height': h['height'], 'waveform': h['waveform']} 
                           for m, h in heights_by_marker.items()],
            'mode': mode,
            'all_end_times': end_times_by_marker
        }

class DefaultEventParser:
    """Parser for default mode (Button Up)"""
    
    def extract_marker(self, line):
        """Extract marker number from log line"""
        match1 = re.search(r'EPDC\]\[(\d+)\]', line)
        if match1:
            return match1.group(1)
        
        match2 = re.search(r'mxc_epdc_fb: \[(\d+)\]', line)
        if match2:
            return match2.group(1)
        
        return None

    def extract_height_and_waveform(self, line):
        """Extract height and waveform information from log line"""
        height_match = re.search(r'height=(\d+)', line)
        if not height_match:
            height_match = re.search(r'width=\d+, height=(\d+)', line)
        
        waveform_patterns = [
            r'new waveform = (?:0x)?[\da-f]+ \(([\w_() ]+)\)', 
            r'waveform:(?:0x)?[\da-f]+ \(([\w_() ]+)\)',
            r'waveform=(?:0x)?[\da-f]+ \(([\w_() ]+)\)',
            r'Sending update\. waveform:(?:0x)?[\da-f]+ \(([\w_() ]+)\)'
        ]
        
        waveform_name = None
        for pattern in waveform_patterns:
            match = re.search(pattern, line)
            if match:
                waveform_name = match.group(1).strip()
                break
        
        if height_match:
            height = int(height_match.group(1))
            return {
                'height': height,
                'waveform': waveform_name if waveform_name else "unknown"
            }
        return None

    def extract_end_timestamp(self, line):
        """Extract end timestamp from log line"""
        match = re.search(r'end time=(\d+)', line)
        if match:
            timestamp_str = match.group(1)
            last_6 = timestamp_str[-6:]
            return int(last_6)
        return None
    
    def extract_start_timestamp(self, line):
        """Extract start timestamp from "button 1 up" event"""
        match = re.search(r'button 1 up (\d+\.\d+)', line)
        if match:
            timestamp_str = match.group(1)
            parts = timestamp_str.split('.')
            if len(parts) == 2:
                last_3_first = parts[0][-3:]
                first_3_second = parts[1][:3]
                result = int(last_3_first + first_3_second)
                return result
        return None

class SwipeEventParser:
    """Parser for swipe mode (Button Down)"""
    
    def extract_marker(self, line):
        """Extract marker number from log line"""
        match1 = re.search(r'EPDC\]\[(\d+)\]', line)
        if match1:
            return match1.group(1)
        
        match2 = re.search(r'mxc_epdc_fb: \[(\d+)\]', line)
        if match2:
            return match2.group(1)
        
        return None

    def extract_height_and_waveform(self, line):
        """Extract height and waveform information from log line"""
        height_match = re.search(r'height=(\d+)', line)
        if not height_match:
            height_match = re.search(r'width=\d+, height=(\d+)', line)
        
        waveform_patterns = [
            r'new waveform = (?:0x)?[\da-f]+ \(([\w_() ]+)\)', 
            r'waveform:(?:0x)?[\da-f]+ \(([\w_() ]+)\)',
            r'waveform=(?:0x)?[\da-f]+ \(([\w_() ]+)\)',
            r'Sending update\. waveform:(?:0x)?[\da-f]+ \(([\w_() ]+)\)'
        ]
        
        waveform_name = None
        for pattern in waveform_patterns:
            match = re.search(pattern, line)
            if match:
                waveform_name = match.group(1).strip()
                break
        
        if height_match:
            height = int(height_match.group(1))
            return {
                'height': height,
                'waveform': waveform_name if waveform_name else "unknown"
            }
        return None

    def extract_end_timestamp(self, line):
        """Extract end timestamp from log line"""
        match = re.search(r'end time=(\d+)', line)
        if match:
            timestamp_str = match.group(1)
            last_6 = timestamp_str[-6:]
            return int(last_6)
        return None
    
    def extract_start_timestamp(self, line):
        """Extract start timestamp from "Sending button 1 down" event"""
        match = re.search(r'Sending button 1 down (\d+\.\d+)', line)
        if match:
            timestamp_str = match.group(1)
            parts = timestamp_str.split('.')
            if len(parts) == 2:
                last_3_first = parts[0][-3:]
                first_3_second = parts[1][:3]
                result = int(last_3_first + first_3_second)
                return result
        return None

class FinalKindleLogAnalyzer(QMainWindow):
    def __init__(self):
        super().__init__()
        self.results = []
        self.current_iteration = 1
        self.all_iterations_data = ""
        self.test_case_title = ""
        self.batch_results = []
        self.loaded_files = []
        self.current_mode = "default"
        self.dark_mode = False
        
        self.setup_ui()
        self.setup_styling()
    
    def setup_ui(self):
        self.setWindowTitle("Final Kindle Log Analyzer - PDF Export & Waveform Boxes")
        self.setGeometry(50, 50, 1600, 1000)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QHBoxLayout(central_widget)
        
        # Create main splitter
        main_splitter = QSplitter(Qt.Horizontal)
        
        # Left panel - Enhanced with all features
        left_panel = self.create_enhanced_left_panel()
        
        # Right panel - Enhanced results with waveform boxes
        right_panel = self.create_enhanced_right_panel()
        
        main_splitter.addWidget(left_panel)
        main_splitter.addWidget(right_panel)
        main_splitter.setSizes([400, 1200])
        
        main_layout.addWidget(main_splitter)
    
    def create_enhanced_left_panel(self):
        """Enhanced left panel with all requested features"""
        panel = QGroupBox("📁 Input & Processing")
        layout = QVBoxLayout()
        
        # Header with dark mode toggle
        header_layout = QHBoxLayout()
        
        title_label = QLabel("Kindle Log Analyzer")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; padding: 10px;")
        title_label.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(title_label)
        
        self.dark_mode_toggle = QCheckBox("Dark Mode")
        self.dark_mode_toggle.toggled.connect(self.toggle_dark_mode)
        header_layout.addWidget(self.dark_mode_toggle)
        
        layout.addLayout(header_layout)
        
        # Test Case and Settings
        settings_group = QGroupBox("🔧 Configuration")
        settings_layout = QVBoxLayout()
        
        # Test case input
        test_case_layout = QHBoxLayout()
        test_case_layout.addWidget(QLabel("Test Case:"))
        self.test_case_input = QLineEdit()
        self.test_case_input.setPlaceholderText("e.g., Kindle_Performance_Test")
        test_case_layout.addWidget(self.test_case_input)
        settings_layout.addLayout(test_case_layout)
        
        # Calculation Mode selection with FIXED suspend
        calc_mode_layout = QHBoxLayout()
        calc_mode_layout.addWidget(QLabel("Mode:"))
        self.calc_mode_combo = QComboBox()
        self.calc_mode_combo.addItems([
            "Default (Button Up)", 
            "Swipe (Button Down)", 
            "Suspend (Power Button - FIXED)"
        ])
        self.calc_mode_combo.currentIndexChanged.connect(self.on_calculation_mode_changed)
        calc_mode_layout.addWidget(self.calc_mode_combo)
        settings_layout.addLayout(calc_mode_layout)
        
        # Processing mode selection
        mode_layout = QHBoxLayout()
        mode_layout.addWidget(QLabel("Processing:"))
        self.processing_mode = QComboBox()
        self.processing_mode.addItems(["Single Entry", "Batch Files"])
        self.processing_mode.currentTextChanged.connect(self.on_processing_mode_changed)
        mode_layout.addWidget(self.processing_mode)
        settings_layout.addLayout(mode_layout)
        
        settings_group.setLayout(settings_layout)
        layout.addWidget(settings_group)
        
        # Single iteration section
        self.single_group = QGroupBox("📝 Single Entry")
        single_layout = QVBoxLayout()
        
        # Log input
        single_layout.addWidget(QLabel("Log Data:"))
        self.log_input = QTextEdit()
        self.log_input.setPlaceholderText("Paste log data here...")
        self.log_input.setMaximumHeight(120)
        single_layout.addWidget(self.log_input)
        
        # Processing buttons
        single_btn_layout = QHBoxLayout()
        self.add_iteration_btn = QPushButton("➕ Add Iteration")
        self.add_iteration_btn.clicked.connect(self.add_iteration)
        self.process_all_btn = QPushButton("🔄 Process All")
        self.process_all_btn.clicked.connect(self.process_all_iterations)
        self.process_all_btn.setEnabled(False)
        
        single_btn_layout.addWidget(self.add_iteration_btn)
        single_btn_layout.addWidget(self.process_all_btn)
        single_layout.addLayout(single_btn_layout)
        
        self.single_group.setLayout(single_layout)
        layout.addWidget(self.single_group)
        
        # Batch processing section
        self.batch_group = QGroupBox("📂 Batch Processing")
        batch_layout = QVBoxLayout()
        
        # File selection
        file_btn_layout = QHBoxLayout()
        self.select_files_btn = QPushButton("🗂️ Select Files")
        self.select_files_btn.clicked.connect(self.select_batch_files)
        self.clear_files_btn = QPushButton("🗑️ Clear")
        self.clear_files_btn.clicked.connect(self.clear_batch_files)
        
        file_btn_layout.addWidget(self.select_files_btn)
        file_btn_layout.addWidget(self.clear_files_btn)
        batch_layout.addLayout(file_btn_layout)
        
        # File list
        self.files_list = QListWidget()
        self.files_list.setMaximumHeight(100)
        batch_layout.addWidget(self.files_list)
        
        # Process button
        self.process_batch_btn = QPushButton("⚡ Process All Files")
        self.process_batch_btn.clicked.connect(self.process_batch_files)
        self.process_batch_btn.setEnabled(False)
        batch_layout.addWidget(self.process_batch_btn)
        
        self.batch_group.setLayout(batch_layout)
        self.batch_group.setVisible(False)
        layout.addWidget(self.batch_group)
        
        # Progress section
        progress_group = QGroupBox("📊 Status")
        progress_layout = QVBoxLayout()
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        progress_layout.addWidget(self.progress_bar)
        
        self.status_label = QLabel("Ready")
        progress_layout.addWidget(self.status_label)
        
        progress_group.setLayout(progress_layout)
        layout.addWidget(progress_group)
        
        # ENHANCED Export section with PDF and improved TXT
        export_group = QGroupBox("💾 Export Options")
        export_layout = QVBoxLayout()
        
        # PDF Export - NEW with highlighting
        self.save_pdf_btn = QPushButton("📑 Generate PDF Report (with Log Highlighting)")
        self.save_pdf_btn.clicked.connect(self.generate_pdf_report)
        self.save_pdf_btn.setEnabled(False)
        export_layout.addWidget(self.save_pdf_btn)
        
        # Enhanced TXT Export with original logs
        self.save_txt_btn = QPushButton("📄 Save Enhanced TXT (with Original Logs)")
        self.save_txt_btn.clicked.connect(self.save_enhanced_txt_report)
        self.save_txt_btn.setEnabled(False)
        export_layout.addWidget(self.save_txt_btn)
        
        # Excel export
        self.export_excel_btn = QPushButton("📊 Export Excel")
        self.export_excel_btn.clicked.connect(self.export_excel_with_highlighting)
        self.export_excel_btn.setEnabled(False)
        export_layout.addWidget(self.export_excel_btn)
        
        # Clear button
        self.clear_all_btn = QPushButton("🗑️ Clear All")
        self.clear_all_btn.clicked.connect(self.clear_all)
        export_layout.addWidget(self.clear_all_btn)
        
        export_group.setLayout(export_layout)
        layout.addWidget(export_group)
        
        layout.addStretch()
        panel.setLayout(layout)
        return panel
    
    def create_enhanced_right_panel(self):
        """Enhanced right panel with waveform boxes and better visualization"""
        panel = QWidget()
        layout = QVBoxLayout()
        
        self.tab_widget = QTabWidget()
        
        # Summary Tab
        self.create_summary_tab()
        
        # Main Results Tab
        self.create_detailed_results_tab()
        
        # NEW: Waveform Boxes Tab - Visual grid layout
        self.create_waveform_boxes_tab()
        
        # Heights/Waveforms Tab
        self.create_heights_waveforms_tab()
        
        # Batch Results Tab
        self.create_batch_results_tab()
        
        layout.addWidget(self.tab_widget)
        panel.setLayout(layout)
        return panel
    
    def create_summary_tab(self):
        """Create summary tab"""
        self.summary_tab = QWidget()
        layout = QVBoxLayout()
        
        self.summary_text = QTextEdit()
        self.summary_text.setReadOnly(True)
        layout.addWidget(self.summary_text)
        
        self.summary_tab.setLayout(layout)
        self.tab_widget.addTab(self.summary_tab, "📊 Summary")
    
    def create_detailed_results_tab(self):
        """Create detailed results tab - COPY-FRIENDLY TABLE"""
        self.results_tab = QWidget()
        layout = QVBoxLayout()
        
        layout.addWidget(QLabel("📋 Main Results (Copy-friendly for Excel)"))
        
        # Main results table - optimized for copying to Excel
        self.results_table = QTableWidget()
        self.results_table.setAlternatingRowColors(True)
        self.results_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.results_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.results_table)
        
        self.results_tab.setLayout(layout)
        self.tab_widget.addTab(self.results_tab, "📋 Main Results")
    
    def create_waveform_boxes_tab(self):
        """Create waveform boxes tab - NEW VISUAL LAYOUT"""
        self.waveform_boxes_tab = QWidget()
        layout = QVBoxLayout()
        
        layout.addWidget(QLabel("📦 Waveform Boxes - Visual Grid Layout"))
        
        # Scrollable area for the boxes
        scroll_area = QScrollArea()
        scroll_widget = QWidget()
        
        # Grid layout for iteration boxes
        self.waveform_grid = QGridLayout(scroll_widget)
        self.waveform_grid.setSpacing(10)
        
        scroll_area.setWidget(scroll_widget)
        scroll_area.setWidgetResizable(True)
        layout.addWidget(scroll_area)
        
        self.waveform_boxes_tab.setLayout(layout)
        self.tab_widget.addTab(self.waveform_boxes_tab, "📦 Waveform Boxes")
    
    def create_heights_waveforms_tab(self):
        """Create heights and waveforms detailed tab"""
        self.heights_tab = QWidget()
        layout = QVBoxLayout()
        
        layout.addWidget(QLabel("📏 All Heights & Waveforms Details"))
        
        # Detailed heights table
        self.heights_table = QTableWidget()
        self.heights_table.setAlternatingRowColors(True)
        self.heights_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        layout.addWidget(self.heights_table)
        
        self.heights_tab.setLayout(layout)
        self.tab_widget.addTab(self.heights_tab, "📏 Heights & Waveforms")
    
    def create_batch_results_tab(self):
        """Create batch results tab"""
        self.batch_tab = QWidget()
        layout = QVBoxLayout()
        
        self.batch_results_text = QTextEdit()
        self.batch_results_text.setReadOnly(True)
        layout.addWidget(self.batch_results_text)
        
        self.batch_tab.setLayout(layout)
        self.tab_widget.addTab(self.batch_tab, "📁 Batch Results")
    
    def create_iteration_waveform_box(self, result):
        """Create a visual box for each iteration's waveform data"""
        box = QFrame()
        box.setFrameStyle(QFrame.StyledPanel)
        box.setStyleSheet(f"""
        QFrame {{
            border: 2px solid {'#0d7377' if not self.dark_mode else '#14a085'};
            border-radius: 8px;
            padding: 10px;
            margin: 5px;
            background-color: {'#ffffff' if not self.dark_mode else '#404040'};
        }}
        QLabel {{
            color: {'#333333' if not self.dark_mode else '#ffffff'};
            font-size: 12px;
        }}
        """)
        
        layout = QVBoxLayout(box)
        
        # Header
        header_label = QLabel(f"🔄 ITERATION_{result['iteration']:02d}")
        header_label.setStyleSheet(f"""
            font-weight: bold; 
            font-size: 14px; 
            color: {'#0d7377' if not self.dark_mode else '#14a085'};
            padding: 5px;
            background-color: {'#f0f8ff' if not self.dark_mode else '#2b2b2b'};
            border-radius: 4px;
        """)
        header_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(header_label)
        
        # Duration - highlighted
        duration_label = QLabel(f"⏱️ Duration: {result['duration']:.3f} seconds")
        duration_label.setStyleSheet(f"""
            font-weight: bold;
            background-color: yellow;
            color: black;
            padding: 3px;
            border-radius: 3px;
        """)
        layout.addWidget(duration_label)
        
        # Selected waveform - highlighted
        selected_label = QLabel(f"🎯 Selected: {result['max_height_waveform']}")
        selected_label.setStyleSheet(f"""
            font-weight: bold; 
            background-color: yellow; 
            color: black;
            padding: 3px;
            border-radius: 3px;
        """)
        layout.addWidget(selected_label)
        
        # Start/Stop info
        times_label = QLabel(f"🔢 {result['start']} → {result['stop']}")
        times_label.setStyleSheet("font-family: monospace; font-size: 11px;")
        layout.addWidget(times_label)
        
        # All heights and waveforms
        layout.addWidget(QLabel("📏 All Heights:"))
        
        for height_info in result['all_heights']:
            is_selected = str(height_info['marker']) == str(result['marker'])
            height_text = f"M{height_info['marker']}: {height_info['height']}px, {height_info['waveform']}"
            
            height_label = QLabel(height_text)
            if is_selected:
                height_label.setStyleSheet("""
                    background-color: yellow; 
                    color: black; 
                    font-weight: bold; 
                    padding: 2px;
                    border-radius: 2px;
                """)
            else:
                height_label.setStyleSheet("font-size: 11px; padding: 1px;")
            
            layout.addWidget(height_label)
        
        # Copy button for this iteration
        copy_btn = QPushButton("📋 Copy Data")
        copy_btn.setMaximumHeight(25)
        copy_btn.clicked.connect(lambda checked, r=result: self.copy_iteration_data(r))
        layout.addWidget(copy_btn)
        
        return box
    
    def copy_iteration_data(self, result):
        """Copy iteration data to clipboard in the requested format"""
        # Create a list to store the height and waveform information

        # Iterate through all heights and extract the required information with numbering
        height_waveform_data = []
        for idx, height_info in enumerate(result['all_heights'], 1):
            height = height_info['height']
            waveform = height_info['waveform']
            height_waveform_data.append(f"{idx}. Height - {height}, Waveform - {waveform}")

        # Join all entries with newlines
        data = "\n".join(height_waveform_data)

        QApplication.clipboard().setText(data)
        self.status_label.setText(f"Copied Iteration {result['iteration']} waveform data to clipboard")
    
    def update_waveform_boxes(self):
        """Update the waveform boxes grid"""
        if not self.results:
            return
        
        # Clear existing boxes
        for i in reversed(range(self.waveform_grid.count())): 
            self.waveform_grid.itemAt(i).widget().setParent(None)
        
        # Add new boxes in a 3-column grid
        cols = 3
        for idx, result in enumerate(self.results):
            row = idx // cols
            col = idx % cols
            
            box = self.create_iteration_waveform_box(result)
            self.waveform_grid.addWidget(box, row, col)
        
        # Add stretch to fill remaining space
        self.waveform_grid.setRowStretch(len(self.results) // cols + 1, 1)
    
    def toggle_dark_mode(self, checked):
        """Toggle between dark and light mode"""
        self.dark_mode = checked
        self.setup_styling()
        # Update waveform boxes with new styling
        if self.results:
            self.update_waveform_boxes()
    
    def setup_styling(self):
        """Setup styling with dark mode support"""
        if self.dark_mode:
            # Dark mode styling
            self.setStyleSheet("""
                QMainWindow {
                    background-color: #2b2b2b;
                    color: #ffffff;
                }
                QWidget {
                    background-color: #2b2b2b;
                    color: #ffffff;
                }
                QGroupBox {
                    font-weight: bold;
                    border: 2px solid #555555;
                    border-radius: 8px;
                    margin: 8px 0px;
                    padding-top: 10px;
                    background-color: #3c3c3c;
                    color: #ffffff;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    left: 10px;
                    padding: 0 8px 0 8px;
                    color: #ffffff;
                }
                QPushButton {
                    background-color: #0d7377;
                    color: white;
                    border: none;
                    padding: 10px 16px;
                    border-radius: 6px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #14a085;
                }
                QPushButton:pressed {
                    background-color: #0a5d61;
                }
                QPushButton:disabled {
                    background-color: #555555;
                    color: #888888;
                }
                QLineEdit, QTextEdit, QComboBox {
                    border: 2px solid #555555;
                    border-radius: 4px;
                    padding: 8px;
                    background-color: #404040;
                    color: #ffffff;
                }
                QLineEdit:focus, QTextEdit:focus, QComboBox:focus {
                    border-color: #0d7377;
                }
                QTableWidget {
                    background-color: #404040;
                    alternate-background-color: #4a4a4a;
                    color: #ffffff;
                    gridline-color: #555555;
                    selection-background-color: #0d7377;
                    selection-color: #ffffff;
                }
                QHeaderView::section {
                    background-color: #0d7377;
                    color: white;
                    padding: 8px;
                    border: 1px solid #555555;
                    font-weight: bold;
                }
                QTabWidget::pane {
                    border: 1px solid #555555;
                    background-color: #3c3c3c;
                }
                QTabBar::tab {
                    background: #505050;
                    color: #ffffff;
                    padding: 10px 16px;
                    margin-right: 2px;
                    border-top-left-radius: 4px;
                    border-top-right-radius: 4px;
                }
                QTabBar::tab:selected {
                    background: #0d7377;
                    color: white;
                }
                QListWidget {
                    background-color: #404040;
                    color: #ffffff;
                    border: 2px solid #555555;
                }
                QProgressBar {
                    border: 2px solid #555555;
                    border-radius: 5px;
                    background-color: #404040;
                }
                QProgressBar::chunk {
                    background-color: #0d7377;
                    border-radius: 3px;
                }
                QScrollArea {
                    background-color: #3c3c3c;
                    border: 1px solid #555555;
                }
                QLabel {
                    color: #ffffff;
                }
            """)
        else:
            # Light mode styling
            self.setStyleSheet("""
                QMainWindow {
                    background-color: #f0f0f0;
                    color: #333333;
                }
                QWidget {
                    background-color: #f0f0f0;
                    color: #333333;
                }
                QGroupBox {
                    font-weight: bold;
                    border: 2px solid #cccccc;
                    border-radius: 8px;
                    margin: 8px 0px;
                    padding-top: 10px;
                    background-color: #ffffff;
                    color: #333333;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    left: 10px;
                    padding: 0 8px 0 8px;
                    color: #333333;
                }
                QPushButton {
                    background-color: #4a90e2;
                    color: white;
                    border: none;
                    padding: 10px 16px;
                    border-radius: 6px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #357abd;
                }
                QPushButton:pressed {
                    background-color: #2968a3;
                }
                QPushButton:disabled {
                    background-color: #cccccc;
                    color: #666666;
                }
                QLineEdit, QTextEdit, QComboBox {
                    border: 2px solid #cccccc;
                    border-radius: 4px;
                    padding: 8px;
                    background-color: #ffffff;
                    color: #333333;
                }
                QLineEdit:focus, QTextEdit:focus, QComboBox:focus {
                    border-color: #4a90e2;
                }
                QTableWidget {
                    background-color: #ffffff;
                    alternate-background-color: #f8f9fa;
                    color: #333333;
                    gridline-color: #e1e8ed;
                    selection-background-color: #4a90e2;
                    selection-color: #ffffff;
                }
                QHeaderView::section {
                    background-color: #4a90e2;
                    color: white;
                    padding: 8px;
                    border: 1px solid #cccccc;
                    font-weight: bold;
                }
                QTabWidget::pane {
                    border: 1px solid #cccccc;
                    background-color: #ffffff;
                    border-radius: 4px;
                }
                QTabBar::tab {
                    background: #e0e0e0;
                    color: #333333;
                    padding: 10px 16px;
                    margin-right: 2px;
                    border-top-left-radius: 4px;
                    border-top-right-radius: 4px;
                }
                QTabBar::tab:selected {
                    background: #4a90e2;
                    color: white;
                }
                QListWidget {
                    background-color: #ffffff;
                    color: #333333;
                    border: 2px solid #cccccc;
                }
                QProgressBar {
                    border: 2px solid #cccccc;
                    border-radius: 5px;
                    background-color: #ffffff;
                }
                QProgressBar::chunk {
                    background-color: #4a90e2;
                    border-radius: 3px;
                }
                QScrollArea {
                    background-color: #ffffff;
                    border: 1px solid #cccccc;
                }
                QLabel {
                    color: #333333;
                }
            """)
    
    def on_calculation_mode_changed(self):
        """Handle calculation mode change"""
        mode_map = {0: "default", 1: "swipe", 2: "suspend"}
        self.current_mode = mode_map.get(self.calc_mode_combo.currentIndex(), "default")
        self.status_label.setText(f"Mode: {self.calc_mode_combo.currentText()}")
    
    def on_processing_mode_changed(self, mode):
        """Handle processing mode change"""
        if mode == "Single Entry":
            self.single_group.setVisible(True)
            self.batch_group.setVisible(False)
        else:
            self.single_group.setVisible(False)
            self.batch_group.setVisible(True)
    
    def add_iteration(self):
        """Add iteration data"""
        log_content = self.log_input.toPlainText().strip()
        if not log_content:
            QMessageBox.warning(self, "Warning", "Please enter log data")
            return
        
        iteration_header = f"\nITERATION_{self.current_iteration:02d}\n"
        self.all_iterations_data += iteration_header + log_content + "\n"
        
        self.current_iteration += 1
        self.log_input.clear()
        self.process_all_btn.setEnabled(True)
        
        self.status_label.setText(f"Added iteration {self.current_iteration-1}. Ready for next iteration.")
    
    def process_all_iterations(self):
        """Process all iterations"""
        if not self.all_iterations_data:
            QMessageBox.warning(self, "Warning", "No iterations to process")
            return
        
        self.progress_bar.setVisible(True)
        self.status_label.setText("Processing iterations...")
        
        # Create and start log processor thread
        self.log_processor = LogProcessor(self.all_iterations_data, self.current_mode)
        self.log_processor.progress_updated.connect(self.progress_bar.setValue)
        self.log_processor.result_ready.connect(self.on_processing_complete)
        self.log_processor.error_occurred.connect(self.on_processing_error)
        self.log_processor.start()
    
    def on_processing_complete(self, data):
        """Handle processing completion"""
        self.results = data['results']
        self.progress_bar.setVisible(False)
        self.update_all_displays()
        self.enable_export_buttons()
        self.status_label.setText(f"Processed {len(self.results)} iterations successfully")
    
    def on_processing_error(self, error):
        """Handle processing error"""
        self.progress_bar.setVisible(False)
        QMessageBox.critical(self, "Processing Error", f"Error during processing: {error}")
        self.status_label.setText("Processing failed")
    
    def update_all_displays(self):
        """Update all result displays"""
        if not self.results:
            return
        
        # Update summary
        self.update_summary_display()
        
        # Update main results table
        self.update_results_table()
        
        # Update waveform boxes - NEW
        self.update_waveform_boxes()
        
        # Update heights and waveforms table
        self.update_heights_table()
    
    def update_summary_display(self):
        """Update summary display"""
        if not self.results:
            return
        
        total_iterations = len(self.results)
        durations = [r['duration'] for r in self.results]
        avg_duration = sum(durations) / total_iterations
        min_duration = min(durations)
        max_duration = max(durations)
        
        summary_html = f"""
        <h2>📊 Processing Summary</h2>
        <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse;">
        <tr><td><b>Test Case:</b></td><td>{self.test_case_input.text() or 'Not specified'}</td></tr>
        <tr><td><b>Processing Mode:</b></td><td>{self.calc_mode_combo.currentText()}</td></tr>
        <tr><td><b>Total Iterations:</b></td><td>{total_iterations}</td></tr>
        <tr><td><b>Average Duration:</b></td><td style="background-color: yellow;">{avg_duration:.3f} seconds</td></tr>
        <tr><td><b>Min Duration:</b></td><td>{min_duration:.3f} seconds</td></tr>
        <tr><td><b>Max Duration:</b></td><td>{max_duration:.3f} seconds</td></tr>
        <tr><td><b>Processing Time:</b></td><td>{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</td></tr>
        </table>

        <h3>🎯 Quick Copy Summary for Excel:</h3>
        <pre style="background-color: #f5f5f5; padding: 10px;">
Iteration    Duration(seconds)    Start   Stop    Height  Waveform
"""

        for result in self.results:
            summary_html += f"{result['iteration']}\t{result['duration']:.3f}\t{result['start']}\t{result['stop']}\t{result['max_height']}\t{result['max_height_waveform']}\n"
        
        summary_html += "</pre>"
        
        self.summary_text.setHtml(summary_html)
    
    def update_results_table(self):
        """Update main results table - optimized for copying"""
        if not self.results:
            return
        
        self.results_table.setRowCount(len(self.results))
        self.results_table.setColumnCount(8)
        
        headers = ['Iteration', 'Duration (seconds)', 'Start Time', 'Stop Time',
                   'Marker', 'Height', 'Selected Waveform', 'Mode']
        self.results_table.setHorizontalHeaderLabels(headers)
        
        for i, result in enumerate(self.results):
            self.results_table.setItem(i, 0, QTableWidgetItem(str(result['iteration'])))

            # Duration is already in seconds
            duration_item = QTableWidgetItem(f"{result['duration']:.3f}")
            duration_item.setBackground(QBrush(QColor(255, 255, 0, 100)))  # Yellow highlighting
            self.results_table.setItem(i, 1, duration_item)
            
            self.results_table.setItem(i, 2, QTableWidgetItem(str(result['start'])))
            self.results_table.setItem(i, 3, QTableWidgetItem(str(result['stop'])))
            self.results_table.setItem(i, 4, QTableWidgetItem(str(result['marker'])))
            self.results_table.setItem(i, 5, QTableWidgetItem(str(result['max_height'])))
            
            # Highlight selected waveform
            waveform_item = QTableWidgetItem(result['max_height_waveform'])
            waveform_item.setBackground(QBrush(QColor(255, 255, 0, 100)))  # Yellow highlighting
            self.results_table.setItem(i, 6, waveform_item)
            
            self.results_table.setItem(i, 7, QTableWidgetItem(result['mode']))
        
        self.results_table.resizeColumnsToContents()
    
    def update_heights_table(self):
        """Update detailed heights and waveforms table"""
        if not self.results:
            return
        
        # Count total rows needed
        total_rows = sum(len(result['all_heights']) for result in self.results)
        
        self.heights_table.setRowCount(total_rows)
        self.heights_table.setColumnCount(6)
        
        headers = ['Iteration', 'Marker', 'Height', 'Waveform', 'Selected', 'End Time']
        self.heights_table.setHorizontalHeaderLabels(headers)
        
        row = 0
        for result in self.results:
            for height_info in result['all_heights']:
                self.heights_table.setItem(row, 0, QTableWidgetItem(str(result['iteration'])))
                self.heights_table.setItem(row, 1, QTableWidgetItem(str(height_info['marker'])))
                self.heights_table.setItem(row, 2, QTableWidgetItem(str(height_info['height'])))
                self.heights_table.setItem(row, 3, QTableWidgetItem(height_info['waveform']))
                
                # Mark if this is the selected marker for final calculation
                is_selected = str(height_info['marker']) == str(result['marker'])
                selected_item = QTableWidgetItem("✓" if is_selected else "")
                if is_selected:
                    selected_item.setBackground(QBrush(QColor(255, 255, 0, 150)))  # Yellow highlighting
                self.heights_table.setItem(row, 4, selected_item)
                
                # Show end time if available
                end_time = ""
                if 'all_end_times' in result and str(height_info['marker']) in result['all_end_times']:
                    end_time = str(result['all_end_times'][str(height_info['marker'])]['time'])
                self.heights_table.setItem(row, 5, QTableWidgetItem(end_time))
                
                row += 1
        
        self.heights_table.resizeColumnsToContents()
    
    def generate_pdf_report(self):
        """Generate enhanced PDF report with highlighting"""
        # Check if we have results (single mode) or batch results
        if not self.results and not self.batch_results:
            QMessageBox.warning(self, "Warning", "No results to generate PDF")
            return

        if not ENHANCED_EXPORTS_AVAILABLE:
            QMessageBox.warning(self, "Warning", "Enhanced PDF export modules not available")
            return

        # Get filename from user
        default_filename = f"{self.test_case_input.text() or 'kindle_analysis'}_report.pdf"
        filename, _ = QFileDialog.getSaveFileName(
            self, "Save PDF Report", default_filename, "PDF Files (*.pdf)"
        )

        if filename:
            try:
                # Create PDF exporter and generate report
                exporter = PdfExporter()

                # Map GUI mode to internal mode
                mode_map = {"Default (Button Up)": "default",
                           "Swipe (Button Down)": "swipe",
                           "Suspend (Power Button - FIXED)": "suspend"}
                current_mode = mode_map.get(self.calc_mode_combo.currentText(), "default")

                # Determine which data to use
                if self.results:
                    # Single entry mode
                    success, message = exporter.generate_pdf_report(
                        self.results, filename, current_mode
                    )
                else:
                    # Batch mode - flatten batch results
                    flattened_results = []
                    for batch in self.batch_results:
                        flattened_results.extend(batch['results'])
                    success, message = exporter.generate_pdf_report(
                        flattened_results, filename, current_mode
                    )

                if success:
                    QMessageBox.information(self, "Success",
                        f"PDF report generated successfully!\n\n"
                        f"Features included:\n"
                        f"• Table of contents with iteration summary\n"
                        f"• Highlighted start/stop points in logs\n"
                        f"• Original log content for each iteration\n"
                        f"• Calculation details and waveform logic\n\n"
                        f"Saved to: {filename}")
                else:
                    QMessageBox.critical(self, "Error", f"Failed to generate PDF: {message}")

            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to generate PDF report: {str(e)}")

    def save_enhanced_txt_report(self):
        """Save enhanced TXT report with original logs for each iteration"""
        # Check if we have results (single mode) or batch results
        if not self.results and not self.batch_results:
            QMessageBox.warning(self, "Warning", "No results to save")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Save Enhanced TXT Report",
            f"{self.test_case_input.text() or 'kindle_analysis'}_enhanced.txt",
            "Text Files (*.txt)"
        )

        if filename:
            try:
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write("=" * 100 + "\n")
                    f.write("KINDLE LOG ANALYZER - ENHANCED REPORT WITH ORIGINAL LOGS\n")
                    f.write("=" * 100 + "\n\n")

                    # Write test case info
                    f.write(f"Test Case: {self.test_case_input.text() or 'Not specified'}\n")
                    f.write(f"Processing Mode: {self.calc_mode_combo.currentText()}\n")
                    f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

                    # Determine which data to use and write appropriate summary
                    if self.results:
                        # Single entry mode
                        f.write(f"Total Iterations: {len(self.results)}\n\n")

                        # Summary statistics for single mode
                        durations = [r['duration'] for r in self.results]
                        f.write("SUMMARY STATISTICS:\n")
                        f.write("-" * 50 + "\n")
                        f.write(f"Average Duration: {sum(durations)/len(durations)/1000:.3f} seconds\n")
                        f.write(f"Min Duration: {min(durations)/1000:.3f} seconds\n")
                        f.write(f"Max Duration: {max(durations)/1000:.3f} seconds\n\n")
                    else:
                        # Batch mode
                        total_files = len(self.batch_results)
                        total_iterations = sum(len(batch['results']) for batch in self.batch_results)
                        f.write(f"Total Files Processed: {total_files}\n")
                        f.write(f"Total Iterations: {total_iterations}\n\n")

                        # Summary statistics for batch mode
                        all_durations = []
                        for batch in self.batch_results:
                            for result in batch['results']:
                                all_durations.append(result['duration'])

                        if all_durations:
                            f.write("SUMMARY STATISTICS:\n")
                            f.write("-" * 50 + "\n")
                            f.write(f"Average Duration: {sum(all_durations)/1000/len(all_durations):.3f} seconds\n")
                            f.write(f"Min Duration: {min(all_durations)/1000:.3f} seconds\n")
                            f.write(f"Max Duration: {max(all_durations)/1000:.3f} seconds\n\n")

                    # Quick copy table
                    f.write("QUICK COPY TABLE (Copy-friendly for Excel):\n")
                    f.write("-" * 80 + "\n")
                    f.write("Iteration\tDuration(sec)\tStart\tStop\tMarker\tHeight\tWaveform\tMode\n")
                    f.write("-" * 80 + "\n")

                    # Write data based on mode
                    if self.results:
                        # Single entry mode
                        for result in self.results:
                            f.write(f"{result['iteration']}\t{result['duration']}\t{result['start']}\t"
                                   f"{result['stop']}\t{result['marker']}\t{result['max_height']}\t"
                                   f"{result['max_height_waveform']}\t{result['mode']}\n")
                    else:
                        # Batch mode
                        for batch in self.batch_results:
                            for result in batch['results']:
                                f.write(f"{batch['filename']}_iter{result['iteration']}\t{result['duration']}\t{result['start']}\t"
                                       f"{result['stop']}\t{result['marker']}\t{result['max_height']}\t"
                                       f"{result['max_height_waveform']}\t{result['mode']}\n")

                    f.write("\n" + "=" * 100 + "\n")
                    f.write("DETAILED ITERATION ANALYSIS WITH ORIGINAL LOGS\n")
                    f.write("=" * 100 + "\n\n")

                    # Write detailed analysis based on mode
                    if self.results:
                        # Single entry mode
                        for result in self.results:
                            f.write(f"ITERATION_{result['iteration']:02d}\n")
                            f.write("=" * 50 + "\n\n")

                            # Original log content
                            f.write("ORIGINAL LOG DATA:\n")
                            f.write("-" * 30 + "\n")

                            if 'original_log' in result:
                                log_lines = result['original_log'].split('\n')
                                for line in log_lines:
                                    if line.strip():
                                        # Mark the start line with highlighting annotation
                                        if 'start_line' in result and result['start_line'] in line:
                                            f.write(f">>> START POINT >>> {line}\n")
                                        else:
                                            f.write(f"{line}\n")
                            else:
                                f.write("Original log data not available\n")

                            f.write("\n")
                            f.write("CALCULATION RESULTS:\n")
                            f.write("-" * 30 + "\n")
                            f.write(f"Start Time: {result['start']}\n")
                            f.write(f"Stop Time: {result['stop']}\n")
                            f.write(f"Duration: {result['duration'] / 1000:.3f} seconds [SELECTED]\n")
                            f.write(f"Selected Marker: {result['marker']}\n")
                            f.write(f"Selected Height: {result['max_height']}px\n")
                            f.write(f"Selected Waveform: {result['max_height_waveform']} [HIGHLIGHTED]\n\n")

                            f.write("ALL HEIGHTS & WAVEFORMS:\n")
                            f.write("-" * 30 + "\n")
                            for height_info in result['all_heights']:
                                marker = height_info['marker']
                                selected = " [SELECTED FOR CALCULATION]" if str(marker) == str(result['marker']) else ""
                                f.write(f"  Marker {marker}: {height_info['height']}px, {height_info['waveform']}{selected}\n")

                            if 'all_end_times' in result:
                                f.write("\nEND TIMES BY MARKER:\n")
                                f.write("-" * 30 + "\n")
                                for marker, end_info in result['all_end_times'].items():
                                    selected = " [SELECTED]" if str(marker) == str(result['marker']) else ""
                                    f.write(f"  Marker {marker}: {end_info['time']}{selected}\n")

                            f.write("\n" + "=" * 50 + "\n\n")
                    else:
                        # Batch mode
                        for batch in self.batch_results:
                            f.write(f"FILE: {batch['filename']}\n")
                            f.write("=" * 60 + "\n\n")

                            if batch['results']:
                                for result in batch['results']:
                                    f.write(f"  ITERATION_{result['iteration']:02d}\n")
                                    f.write("  " + "-" * 40 + "\n\n")

                                    # Original log content
                                    f.write("  ORIGINAL LOG DATA:\n")
                                    f.write("  " + "-" * 20 + "\n")

                                    if 'original_log' in result:
                                        log_lines = result['original_log'].split('\n')
                                        for line in log_lines:
                                            if line.strip():
                                                # Mark the start line with highlighting annotation
                                                if 'start_line' in result and result['start_line'] in line:
                                                    f.write(f"  >>> START POINT >>> {line}\n")
                                                else:
                                                    f.write(f"  {line}\n")
                                    else:
                                        f.write("  Original log data not available\n")

                                    f.write("\n")
                                    f.write("  CALCULATION RESULTS:\n")
                                    f.write("  " + "-" * 20 + "\n")
                                    f.write(f"  Start Time: {result['start']}\n")
                                    f.write(f"  Stop Time: {result['stop']}\n")
                                    f.write(f"  Duration: {result['duration'] / 1000:.3f} seconds [SELECTED]\n")
                                    f.write(f"  Selected Marker: {result['marker']}\n")
                                    f.write(f"  Selected Height: {result['max_height']}px\n")
                                    f.write(f"  Selected Waveform: {result['max_height_waveform']} [HIGHLIGHTED]\n\n")

                                    f.write("  ALL HEIGHTS & WAVEFORMS:\n")
                                    f.write("  " + "-" * 20 + "\n")
                                    for height_info in result['all_heights']:
                                        marker = height_info['marker']
                                        selected = " [SELECTED FOR CALCULATION]" if str(marker) == str(result['marker']) else ""
                                        f.write(f"    Marker {marker}: {height_info['height']}px, {height_info['waveform']}{selected}\n")

                                    if 'all_end_times' in result:
                                        f.write("\n  END TIMES BY MARKER:\n")
                                        f.write("  " + "-" * 20 + "\n")
                                        for marker, end_info in result['all_end_times'].items():
                                            selected = " [SELECTED]" if str(marker) == str(result['marker']) else ""
                                            f.write(f"    Marker {marker}: {end_info['time']}{selected}\n")

                                    f.write("\n  " + "=" * 40 + "\n\n")
                            else:
                                f.write("  No valid results found in this file.\n\n")

                    f.write("=" * 100 + "\n")
                    f.write("END OF ENHANCED REPORT\n")
                    f.write("=" * 100 + "\n")

                QMessageBox.information(self, "Success", f"Enhanced TXT report with original logs saved to:\n{filename}")

            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save enhanced TXT report: {str(e)}")
    
    def export_excel_with_highlighting(self):
        """Export to Excel with yellow highlighting"""
        # Check if we have results (single mode) or batch results
        if not self.results and not self.batch_results:
            QMessageBox.warning(self, "Warning", "No results to export")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Export Excel with Highlighting",
            f"{self.test_case_input.text() or 'kindle_analysis'}_highlighted.xlsx",
            "Excel Files (*.xlsx)"
        )

        if filename:
            try:
                workbook = openpyxl.Workbook()

                # Main Results Sheet
                sheet = workbook.active
                sheet.title = "Main Results"

                # Headers
                headers = ['Iteration', 'Duration (seconds)', 'Start Time', 'Stop Time',
                          'Marker', 'Height', 'Selected Waveform', 'Mode']
                for col, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")

                # Data with highlighting
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                # Write data based on mode
                row = 2
                if self.results:
                    # Single entry mode
                    for result in self.results:
                        sheet.cell(row=row, column=1, value=result['iteration'])

                        # Highlight duration
                        duration_cell = sheet.cell(row=row, column=2, value=result['duration'])
                        duration_cell.fill = yellow_fill

                        sheet.cell(row=row, column=3, value=result['start'])
                        sheet.cell(row=row, column=4, value=result['stop'])
                        sheet.cell(row=row, column=5, value=result['marker'])
                        sheet.cell(row=row, column=6, value=result['max_height'])

                        # Highlight selected waveform
                        waveform_cell = sheet.cell(row=row, column=7, value=result['max_height_waveform'])
                        waveform_cell.fill = yellow_fill

                        sheet.cell(row=row, column=8, value=result['mode'])
                        row += 1
                else:
                    # Batch mode
                    for batch in self.batch_results:
                        for result in batch['results']:
                            sheet.cell(row=row, column=1, value=f"{batch['filename']}_iter{result['iteration']}")

                            # Highlight duration
                            duration_cell = sheet.cell(row=row, column=2, value=result['duration'])
                            duration_cell.fill = yellow_fill

                            sheet.cell(row=row, column=3, value=result['start'])
                            sheet.cell(row=row, column=4, value=result['stop'])
                            sheet.cell(row=row, column=5, value=result['marker'])
                            sheet.cell(row=row, column=6, value=result['max_height'])

                            # Highlight selected waveform
                            waveform_cell = sheet.cell(row=row, column=7, value=result['max_height_waveform'])
                            waveform_cell.fill = yellow_fill

                            sheet.cell(row=row, column=8, value=result['mode'])
                            row += 1

                # Auto-size columns
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                length = len(str(cell.value))
                                if length > max_length:
                                    max_length = length
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width

                # Detailed Heights Sheet
                detail_sheet = workbook.create_sheet(title="Heights & Waveforms")

                detail_headers = ['Iteration', 'Marker', 'Height', 'Waveform', 'Selected', 'End Time']
                for col, header in enumerate(detail_headers, 1):
                    cell = detail_sheet.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")

                detail_row = 2
                # Write detailed data based on mode
                if self.results:
                    # Single entry mode
                    for result in self.results:
                        for height_info in result['all_heights']:
                            detail_sheet.cell(row=detail_row, column=1, value=result['iteration'])
                            detail_sheet.cell(row=detail_row, column=2, value=height_info['marker'])
                            detail_sheet.cell(row=detail_row, column=3, value=height_info['height'])
                            detail_sheet.cell(row=detail_row, column=4, value=height_info['waveform'])

                            # Highlight selected rows
                            is_selected = str(height_info['marker']) == str(result['marker'])
                            selected_cell = detail_sheet.cell(row=detail_row, column=5, value="✓" if is_selected else "")
                            if is_selected:
                                selected_cell.fill = yellow_fill
                                # Highlight entire row for selected marker
                                for col in range(1, 7):
                                    detail_sheet.cell(row=detail_row, column=col).fill = yellow_fill

                            # End time
                            end_time = ""
                            if 'all_end_times' in result and str(height_info['marker']) in result['all_end_times']:
                                end_time = result['all_end_times'][str(height_info['marker'])]['time']
                            detail_sheet.cell(row=detail_row, column=6, value=end_time)

                            detail_row += 1
                else:
                    # Batch mode
                    for batch in self.batch_results:
                        for result in batch['results']:
                            for height_info in result['all_heights']:
                                detail_sheet.cell(row=detail_row, column=1, value=f"{batch['filename']}_iter{result['iteration']}")
                                detail_sheet.cell(row=detail_row, column=2, value=height_info['marker'])
                                detail_sheet.cell(row=detail_row, column=3, value=height_info['height'])
                                detail_sheet.cell(row=detail_row, column=4, value=height_info['waveform'])

                                # Highlight selected rows
                                is_selected = str(height_info['marker']) == str(result['marker'])
                                selected_cell = detail_sheet.cell(row=detail_row, column=5, value="✓" if is_selected else "")
                                if is_selected:
                                    selected_cell.fill = yellow_fill
                                    # Highlight entire row for selected marker
                                    for col in range(1, 7):
                                        detail_sheet.cell(row=detail_row, column=col).fill = yellow_fill

                                # End time
                                end_time = ""
                                if 'all_end_times' in result and str(height_info['marker']) in result['all_end_times']:
                                    end_time = result['all_end_times'][str(height_info['marker'])]['time']
                                detail_sheet.cell(row=detail_row, column=6, value=end_time)

                                detail_row += 1

                # Auto-size columns for detail sheet
                for column in detail_sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                length = len(str(cell.value))
                                if length > max_length:
                                    max_length = length
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    detail_sheet.column_dimensions[column_letter].width = adjusted_width

                workbook.save(filename)
                QMessageBox.information(self, "Success", f"Excel file with highlighting saved to:\n{filename}")

            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save Excel file: {str(e)}")
                    
                workbook.save(filename)
                QMessageBox.information(self, "Success", f"Excel file with highlighting saved to:\n{filename}")

            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save Excel file: {str(e)}")
    
    def select_batch_files(self):
        """Select files for batch processing"""
        files, _ = QFileDialog.getOpenFileNames(
            self, "Select Log Files", "", 
            "Log Files (*.log *.txt);;All Files (*)"
        )
        if files:
            self.loaded_files = files
            self.files_list.clear()
            for file in files:
                self.files_list.addItem(os.path.basename(file))
            self.process_batch_btn.setEnabled(True)
    
    def clear_batch_files(self):
        """Clear selected batch files"""
        self.loaded_files = []
        self.files_list.clear()
        self.process_batch_btn.setEnabled(False)
    
    def process_batch_files(self):
        """Process batch files"""
        if not self.loaded_files:
            return

        self.status_label.setText("Processing batch files...")
        self.batch_results = []

        for file_path in self.loaded_files:
            try:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()

                # Process content using log processor
                processor = LogProcessor(content, self.current_mode)
                # Simulate processing for batch
                iterations = re.split(r'ITERATION_(\d+)', content)[1:]
                if not iterations:
                    iterations = ["01", content]

                iteration_pairs = []
                for i in range(0, len(iterations), 2):
                    if i+1 < len(iterations):
                        iteration_num = iterations[i]
                        iteration_content = iterations[i+1]
                        iteration_pairs.append((iteration_num, iteration_content))

                file_results = []

                for iteration_num, iteration_content in iteration_pairs:
                    lines = iteration_content.split('\n')
                    result = processor.process_iteration(lines, iteration_num, self.current_mode)
                    if result:
                        result['original_log'] = iteration_content.strip()
                        file_results.append(result)

                self.batch_results.append({
                    'filename': os.path.basename(file_path),
                    'results': file_results
                })

            except Exception as e:
                QMessageBox.warning(self, "Warning", f"Error processing {file_path}: {str(e)}")

        self.update_batch_display()
        if self.batch_results:
            self.enable_export_buttons()
        self.status_label.setText(f"Processed {len(self.loaded_files)} files")

    def update_batch_display(self):
        """Update batch results display"""
        if not self.batch_results:
            return
        
        batch_html = "<h2>📁 Batch Processing Results</h2>"
        
        for batch in self.batch_results:
            batch_html += f"<h3>📄 {batch['filename']}</h3>"
            if batch['results']:
                batch_html += "<table border='1' cellpadding='5' cellspacing='0'>"
                batch_html += "<tr><th>Iteration</th><th>Duration</th><th>Start</th><th>Stop</th><th>Height</th><th>Waveform</th></tr>"
                
                for result in batch['results']:
                    batch_html += f"""
                    <tr>
                    <td>{result['iteration']}</td>
                    <td style='background-color: yellow;'>{result['duration']}</td>
                    <td>{result['start']}</td>
                    <td>{result['stop']}</td>
                    <td>{result['max_height']}</td>
                    <td style='background-color: yellow;'>{result['max_height_waveform']}</td>
                    </tr>
                    """
                batch_html += "</table><br>"
            else:
                batch_html += "<p>No valid results found.</p>"
        
        self.batch_results_text.setHtml(batch_html)
    
    def enable_export_buttons(self):
        """Enable export buttons"""
        self.save_pdf_btn.setEnabled(True)
        self.save_txt_btn.setEnabled(True)
        self.export_excel_btn.setEnabled(True)
    
    def clear_all(self):
        """Clear all data"""
        self.results = []
        self.batch_results = []
        self.loaded_files = []
        self.all_iterations_data = ""
        self.current_iteration = 1
        
        self.log_input.clear()
        self.files_list.clear()
        self.summary_text.clear()
        self.results_table.setRowCount(0)
        self.heights_table.setRowCount(0)
        self.batch_results_text.clear()
        
        # Clear waveform boxes
        for i in reversed(range(self.waveform_grid.count())): 
            self.waveform_grid.itemAt(i).widget().setParent(None)
        
        self.save_pdf_btn.setEnabled(False)
        self.save_txt_btn.setEnabled(False)
        self.export_excel_btn.setEnabled(False)
        self.process_all_btn.setEnabled(False)
        self.process_batch_btn.setEnabled(False)
        
        self.status_label.setText("Ready")

# Fixed main execution block - only one instance at the end
if __name__ == '__main__':
    app = QApplication(sys.argv)
    
    # Set application-wide font - use a cross-platform font
    font = QFont("Arial", 10)  # Changed from "Segoe UI" to "Arial"
    app.setFont(font)
    
    window = FinalKindleLogAnalyzer()
    window.show()
    
    sys.exit(app.exec_())